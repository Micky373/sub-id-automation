# Importing useful libraries
import OleFileIO_PL # For reading excel files
import pandas as pd # For dataframe related tasks
from openpyxl.styles import PatternFill, Font , Alignment# For excel formating
from openpyxl.utils import get_column_letter
import swifter # For paralalizing long taking tasks
import zipfile # For zipping file together

# For operating system related tasks
import os 
import shutil

import numpy as np # For numerical tasks
import streamlit as st # For simple web app

# A function to zip the files
def zip_files(file_paths, output_zip_file):
    with zipfile.ZipFile(output_zip_file, 'w') as zipf:
        for file in file_paths:
            zipf.write(file, os.path.basename(file))

# Reading excel file
def read_excel_file(path):
    with open(path,'rb') as file:
        ole = OleFileIO_PL.OleFileIO(file)
        if ole.exists('Workbook'):
            d = ole.openstream('Workbook')
            data=pd.read_excel(d,engine='xlrd')
    return data

# Function to apply formatting to a sheet
def format_sheet(sheet, conditional_formating = False,error_sheet=False,warning_sheet=False,findings=False):
    # Color the first row black
    black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    white_bold_font = Font(color="FFFFFF", bold=True)

    # Example merging cells and adding custom text and styles
    def merge_and_style_cells(start_row, end_row, start_col, end_col, text, fill, font_size):
        sheet.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)
        merged_cell = sheet.cell(row=start_row, column=start_col)
        merged_cell.value = text
        merged_cell.fill = fill
        merged_cell.font = Font(color="000000", size=font_size)
        merged_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Calculate the width of the text to adjust the column width
        text_length = len(text)
        for col in range(start_col, end_col + 1):
            column_letter = get_column_letter(col)
            current_width = sheet.column_dimensions[column_letter].width
            if current_width is None:
                current_width = 0
            # Adjust the column width to be the maximum of current width or text length
            new_width = max(current_width, text_length + 2)
            sheet.column_dimensions[column_letter].width = new_width
    
    if findings != False:
        unique_findings,must_stop_findings,below_40_findings,exactly_40_findings,above_40_findings,errors = findings
        # Merge and style specific rows
        merge_and_style_cells(2, 5, 1, 13, unique_findings, PatternFill(start_color ="FFFF99", end_color ="FFFF99", fill_type ="solid"), 30)
        merge_and_style_cells(6, 9, 1, 13, must_stop_findings, PatternFill(start_color ="FF6666", end_color ="FF6666", fill_type ="solid"), 30)
        merge_and_style_cells(10, 13, 1, 13, below_40_findings, PatternFill(start_color ="FFCCCC", end_color ="FFCCCC", fill_type ="solid"), 30)
        merge_and_style_cells(14, 17, 1, 13, exactly_40_findings, PatternFill(start_color ="FFFF99", end_color ="FFFF99", fill_type ="solid"), 30)
        merge_and_style_cells(18, 21, 1, 13, above_40_findings, PatternFill(start_color ="CCFFCC", end_color ="CCFFCC", fill_type ="solid"), 30)
        merge_and_style_cells(22, 24, 1, 13, errors, PatternFill(start_color = "FF6666", end_color ="CCFFCC", fill_type ="solid"), 30)

    if findings == False:
        for cell in sheet[1]:
            cell.fill = black_fill
            cell.font = white_bold_font # Change font color to white for readability

    # Adjust column widths to fit the contents
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column].width = adjusted_width

    # Conditional formatting based on average percentage
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

    if conditional_formating:

        # Find the column with the header "Average Percentage"
        avg_col = None
        for cell in sheet[1]:
            if cell.value == "Average Percentage":
                avg_col = cell.column
                break

        if avg_col is None:
            print("Column 'Average Percentage' not found.")
            return

        for row in sheet.iter_rows(min_row=2):
            avg_cell = row[avg_col - 1]
            try:
                value = float(avg_cell.value)
                if value < 40:
                    row_fill = red_fill
                elif 40 <= value < 70:
                    row_fill = yellow_fill
                elif value >= 70:
                    row_fill = green_fill
                
                for cell in row:
                    cell.fill = row_fill
            except (ValueError, TypeError):
                continue
    
    if error_sheet:

        for row in sheet.iter_rows(min_row=2):

            for cell in row:
                cell.fill = red_fill 

    if warning_sheet:

        for row in sheet.iter_rows(min_row=2):

            for cell in row:
                cell.fill = yellow_fill

    # Coloring the cell containing the total data
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == 'Total':
                cell.fill = black_fill
                cell.font = white_bold_font


def return_revenue(revenue_data,date,name,reve_id,sub_id):

    temp_rev_data = revenue_data[
        (revenue_data['Company'] == name) & (revenue_data['Affiliate'] == reve_id) & (revenue_data['Lead Date'] == date) & \
            (revenue_data['s1'] == sub_id)]
    
    revenue = temp_rev_data['Revenue'].sum()
   
    return revenue

def return_cost(revenue_data,date,name,reve_id,sub_id):

    temp_rev_data = revenue_data[
        (revenue_data['Company'] == name) & (revenue_data['Affiliate'] == reve_id) & (revenue_data['Lead Date'] == date) & \
            (revenue_data['s1'] == sub_id)]
    
    cost = temp_rev_data['Cost'].sum()
    
    return cost

def apply_with_progress(df, func, progress_bar):
    total_chunks = 10
    chunk_size = len(df) // total_chunks

    def process_chunk(chunk):
        return chunk.swifter.apply(func, axis=1)

    results = []
    for i in range(total_chunks):
        start = i * chunk_size
        end = (i + 1) * chunk_size if i < total_chunks - 1 else len(df)
        chunk = df.iloc[start:end]
        results.append(process_chunk(chunk))
        progress_bar.progress((i + 1) / total_chunks)

    return pd.concat(results)

def get_report(click_reg_path,revenue_data_path,zip_file_path):

    # Creating an empty directory to save all the findings
    temp_dir = 'synthesized_data/temp_dir/'
    if not os.path.exists(temp_dir): os.mkdir(temp_dir)

    # Checking if there is a zipped file and deleting it
    if os.path.exists('synthesized_data/files.zip'):
        os.remove('synthesized_data/files.zip')

    # Reading the data
    df = read_excel_file(click_reg_path)
    revenue_data = read_excel_file(revenue_data_path)

    df['Percentage'] = round((df['User Registration'] / df['Cake Clicks (All Clicks)']) * 100 , 2)

    # So as we can see S2,S3,S4,S5 contains all null values so we don't need those columns
    useful_columns = [column for column in df.columns if column not in ['S2','S3','S4','S5']]
    df = df[useful_columns]

    # Cleaning the data by removing the null values
    df.dropna(inplace=True)

    # Removing the s2= and s2&#61; value from the s1 columns
    df['S1'] = df['S1'].apply(lambda s1:str(s1).replace('s2','').replace('=','').replace('&#61;',''))
    revenue_data['s1'] = revenue_data['s1'].apply(lambda s1:str(s1).replace('s2','').replace('=','').replace('&#61;',''))

    # Create Streamlit progress bars with text
    st.text('Processing revenue calculations...')
    revenue_progress_bar = st.progress(0)
    st.text('Processing cost calculations...')
    cost_progress_bar = st.progress(0)
    st.text('Overall publisher progress...')
    overall_progress_bar = st.progress(0)

    # Getting the revenue and sub id analysis
    # df['Revenue'] = df.swifter.apply(
    #     lambda row: return_revenue(revenue_data, row['Date'], row['Affiliate Name'], row['Revenue Tracker ID'], row['S1']),
    #     axis=1
    # )

    # Getting the revenue and sub id analysis
    df['Revenue'] = apply_with_progress(
        df, lambda row: return_revenue(revenue_data, row['Date'], row['Affiliate Name'], row['Revenue Tracker ID'], row['S1']),
        revenue_progress_bar
    )

    # df['Cost'] = df.swifter.apply(
    #     lambda row: return_cost(revenue_data, row['Date'], row['Affiliate Name'], row['Revenue Tracker ID'], row['S1']),
    #     axis=1
    # )

    df['Cost'] = apply_with_progress(
        df, lambda row: return_cost(revenue_data, row['Date'], row['Affiliate Name'], row['Revenue Tracker ID'], row['S1']),
        cost_progress_bar
    )

    df['Margin'] = (df['Revenue'] - df['Cost']) / df['Revenue']
    df['Margin'] = df['Margin'].apply(lambda margin:round(margin,2) if margin != np.inf else 0)

    # Getting all the publishers
    publishers = [publisher for publisher in list(set(df['Affiliate Name'].values))]

    df['Percentage'] = df['Percentage'].apply(lambda perc:perc if perc != np.inf else 0)

    total_publishers = len(publishers)
    progress = 0

    # Itterating through all the publishers and getting their data    
    for publisher in publishers:

        # Creating a dictionary containing a dictionary to contain all the dfs
        data_frames = {}

        print(f'------Doing the analysis for {publisher}------')

        # Getting the publishers data frame
        df_for_analysis = df[df['Affiliate Name'] == publisher]
        data_frames['Original File'] = df_for_analysis

        # Getting the rows with error ids
        error_df = df_for_analysis[(df_for_analysis['Cake Clicks (All Clicks)'] == 0) & (df_for_analysis['User Registration'] > 0)]
        error_df = error_df.sort_values(by = 'S1',ascending = False)

        # Getting abnormal sub ids
        abnormal_df = df_for_analysis[(df_for_analysis['User Registration'] > df_for_analysis['Cake Clicks (All Clicks)']) & \
                                    (df_for_analysis['Cake Clicks (All Clicks)'] != 0)]

        # Now let us take the clean df that contain correct sub ids
        df_for_analysis = df_for_analysis[(df_for_analysis['Cake Clicks (All Clicks)'] != 0) & \
                                        (df_for_analysis['User Registration'] <= df_for_analysis['Cake Clicks (All Clicks)'])]
        
        # st.dataframe(df_for_analysis)

        # Now let us see how many duplicate sub ids we have
        temp_df = pd.DataFrame(
            df_for_analysis['S1'].value_counts()
        )
        temp_df.reset_index(inplace=True)
        temp_df.rename(
            columns = {
                'index' : 'S1',
                'count' : 'Number of Repetition'
            },
            inplace = True
        )

        # st.dataframe(temp_df)
        duplicated_sub_ids = temp_df[temp_df['Number of Repetition'] > 1]['S1'].values

        # Let us create a new column letting us know if the id is duplicated or not
        df_for_analysis['Duplicate'] = df_for_analysis['S1'].apply(lambda s1 : 'Yes' if s1 in duplicated_sub_ids else 'No')

        # Finding the duplicated and not duplicated dataframe
        dup_df = df_for_analysis[df_for_analysis['Duplicate'] == 'Yes']
        data_frames['Dup_Sub_IDs'] = dup_df
        no_dup_df = df_for_analysis[df_for_analysis['Duplicate'] == 'No']
        data_frames['Not_Dup_Sub_IDs'] = no_dup_df

        # Getting the duplicated sub ids that has more than 40% registration rate
        dup_more_40 = dup_df[(dup_df['Percentage'] > 40) & (dup_df['Percentage'] != np.inf)]
        data_frames['Dup_Sub_IDs>40%_regi_rate'] = dup_more_40

        # Getting the duplicated sub ids that has less than 40% registration rate
        dup_less_40 = dup_df[(dup_df['Percentage'] < 40) & (dup_df['Percentage'] != np.inf)]
        data_frames['Dup_Sub_IDs<40%_regi_rate'] = dup_less_40

        # Getting the duplicated sub ids that has exactly 40% registration rate
        dup_exactly_40 = dup_df[(dup_df['Percentage'] == 40) & (dup_df['Percentage'] != np.inf)]
        data_frames['Dup_Sub_IDs=40%_regi_rate'] = dup_exactly_40

        # Getting the non duplicated sub ids that has more than 40% registration rate
        non_dup_more_40 = no_dup_df[(no_dup_df['Percentage'] > 40) & (no_dup_df['Percentage'] != np.inf)]
        data_frames['Not-Dup_Sub_IDs>40%_reg_rate'] = non_dup_more_40

        # Getting the non duplicated sub ids that has less than 40% registration rate
        non_dup_less_40 = no_dup_df[(no_dup_df['Percentage'] < 40) & (no_dup_df['Percentage'] != np.inf)]
        data_frames['Not-Dup_Sub_IDs<40%_reg_rate'] = non_dup_less_40

        # Getting the non duplicated sub ids that has exactly 40% registration rate
        non_dup_exactly_40 = no_dup_df[(no_dup_df['Percentage'] == 40) & (no_dup_df['Percentage'] != np.inf)]
        data_frames['Not-Dup_Sub_IDs=40%_reg_rate'] = non_dup_exactly_40

        # Making each duplicated sub ids together
        dup_sorted_df = dup_df.sort_values(by=['Duplicate', 'S1'],ascending = False)

        # Getting the average value for each duplicated sub ids
        def get_average_registration_rate(id):

            total_clicks = dup_sorted_df[dup_sorted_df['S1'] == id]['Cake Clicks (All Clicks)'].sum()

            total_registration = dup_sorted_df[dup_sorted_df['S1'] == id]['User Registration'].sum()

            average_reg_rate = round((total_registration/total_clicks) * 100 , 2)

            return total_clicks,total_registration,average_reg_rate
        
        registration_mapping = {}

        reg_mapping = {}

        click_mapping = {}

        for id in duplicated_sub_ids:

            total_clicks,total_registration,avg_reg_rate = get_average_registration_rate(id)
            registration_mapping[id] = avg_reg_rate
            reg_mapping[id] = total_registration
            click_mapping[id] = total_clicks

        # Sorting the sub ids
        dup_sorted_df['Average Percentage'] = dup_sorted_df['S1'].apply(lambda s1:registration_mapping[s1])
        dup_sorted_df = dup_sorted_df.sort_values(by = ['Average Percentage','S1'],ascending = False)
        temp_dfs = []
        temp_list_id = []
        for id in list(dup_sorted_df['S1'].values):

            if id not in temp_list_id:
                temp_dfs.append(dup_sorted_df[dup_sorted_df['S1'] == id])
                total_data = ['','','','','Total',reg_mapping[id],click_mapping[id],'','','','','',registration_mapping[id]]

                total_data_df = pd.DataFrame([total_data],
                                                columns=dup_sorted_df.columns)
                
                temp_dfs.append(total_data_df)

            temp_list_id.append(id)

        if len(temp_dfs) > 0:
            dup_sorted_df = pd.concat(temp_dfs,axis = 0)
        else:
            pass
            
        data_frames['Duplicated_Sorted_by_avg_reg_rate'] = dup_sorted_df

        data_frames['Abnormal Sub IDs'] = abnormal_df
        data_frames['Errors'] = error_df
        data_frames['Must Stop 0%'] = df_for_analysis[df_for_analysis['Percentage'] == 0]
        data_frames['Conclusion'] = pd.DataFrame()

        with pd.ExcelWriter(f'{temp_dir}{publisher}.xlsx', engine='openpyxl') as writer:
            for sheet_name, saving_df in data_frames.items():

                saving_df.to_excel(writer, index = False, sheet_name=sheet_name)

                # Access the workbook and worksheet
                workbook = writer.book
                worksheet = workbook[sheet_name]

                conditional_formating = False
                error_sheet = False
                warning_sheet = False
                findings = False

                # Apply styles
                if sheet_name == 'Duplicated_Sorted_by_avg_reg_rate': conditional_formating = True
                    
                if sheet_name in ['Must Stop 0%','Errors'] : error_sheet = True

                if sheet_name == 'Abnormal Sub IDs' : warning_sheet = True

                if sheet_name == 'Conclusion' : 
                    unique_ids = len(list(set(df_for_analysis['S1'].values)))
                    must_stop = df_for_analysis[df_for_analysis['Percentage'] == 0].shape[0]
                    below_40 = df_for_analysis[df_for_analysis['Percentage'] < 40].shape[0]
                    exact_40 = df_for_analysis[df_for_analysis['Percentage'] == 40].shape[0]
                    abv_40 = df_for_analysis[df_for_analysis['Percentage'] > 40].shape[0]
                    errors = error_df.shape[0]

                    findings = [
                        f'We have {unique_ids} unique s1',
                        f'We have {must_stop} s1 Must stop ids 0% ',
                        f'{below_40} Below 40 %',
                        f'{exact_40} exactly 40%',
                        f'{abv_40} good s1 above 40%',
                        f'{errors} error s1'
                    ]

                format_sheet(worksheet,conditional_formating,error_sheet,warning_sheet,findings)

        progress += 1
        overall_progress_bar.progress(progress / total_publishers)

        print(f'***Finished saving the file***')

    print('////////////////Zipping the files////////////////')
    # Zipping the already created dataframes
    files = [temp_dir + file for file in os.listdir(temp_dir)]
    zip_files(files,zip_file_path)

    # Removing the temp directory
    shutil.rmtree(temp_dir)

